#!/usr/bin/env python3
"""解析 .xlsx 测试报告并导出 Markdown 报告。"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

PROJECT_ROOT = SCRIPTS_DIR.parents[1]
DEFAULT_TEMPLATE = (
    PROJECT_ROOT / "docs" / "retrospective" / "templates" / "xlsx-test-report-template.md"
)
DEFAULT_SUMMARY_TEMPLATE = (
    PROJECT_ROOT
    / "docs"
    / "retrospective"
    / "templates"
    / "release-gate-summary-template.md"
)

METRIC_KEY_MAP = {
    "总用例": "total_cases",
    "所有用例": "total_cases",
    "PASS": "pass",
    "FAIL": "fail",
    "NOTEST": "notest",
    "NOT TEST": "notest",
    "NT": "notest",
    "BLOCK": "block",
    "DI": "di",
    "DI值": "di",
    "严重问题数": "serious_issues",
    "所有严重问题": "serious_issues",
}

STATUS_VALUE_MAP = {
    "PASS": "pass",
    "FAIL": "fail",
    "NT": "notest",
    "NOTEST": "notest",
    "NOT TEST": "notest",
    "BLOCK": "block",
}

RISK_KEYWORDS = {
    "重启恢复": ("重启", "死机", "崩溃", "异常恢复", "软启", "断电恢复"),
    "弱网": ("弱网", "穿墙", "丢包", "重连", "断流", "网络异常"),
    "存储回放": ("TF卡", "TF", "存储", "录像", "回放", "文件损坏"),
    "音频": ("底噪", "回声", "啸叫", "破音", "吞字", "无声", "杂音"),
    "升级稳定性": ("升级失败", "升级后", "版本回退", "升级重启"),
    "预览稳定性": ("卡顿", "丢帧", "花屏", "黑屏", "拉流失败", "拉流", "延迟", "不同步"),
}

RISK_PRIORITY = (
    "重启恢复",
    "弱网",
    "存储回放",
    "音频",
    "升级稳定性",
    "预览稳定性",
)

PLATFORM_SEMANTIC_DEFAULTS = {
    "ha_domain": "tuya",
    "entity_scope": "camera",
}

PLATFORM_RISK_PROFILES = {
    "重启恢复": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "status_range", "diagnostics"),
        "diagnostic_focus": "优先核对设备在线状态恢复、关键 DP 状态刷新与诊断导出中的异常恢复痕迹。",
    },
    "弱网": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "function", "diagnostics"),
        "diagnostic_focus": "结合实体状态刷新延迟、云侧推送连续性与诊断字段，复核弱网下的重连与状态同步风险。",
    },
    "存储回放": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "function", "status_range"),
        "diagnostic_focus": "重点映射录像、回放和存储相关能力字段，确认平台侧是否能观测到存储异常。",
    },
    "音频": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "function", "status_range"),
        "diagnostic_focus": "结合摄像头实体能力与诊断字段，复核音视频相关功能项是否存在异常状态或能力缺口。",
    },
    "升级稳定性": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "diagnostics", "token_info"),
        "diagnostic_focus": "复核升级后的在线状态恢复、云侧令牌续期与设备诊断信息是否保持稳定。",
    },
    "预览稳定性": {
        "integration_traits": ("hub", "cloud_push"),
        "observation_surfaces": ("status", "function", "status_range"),
        "diagnostic_focus": "重点关注实时预览相关实体能力、状态变化和功能范围，确认平台侧是否能反映画面链路异常。",
    },
}


def configure_stdio() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="解析 .xlsx 测试报告并导出 Markdown")
    parser.add_argument("--input", required=True, help="输入 .xlsx 文件路径")
    parser.add_argument("--output", required=True, help="输出 Markdown 文件路径")
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="Markdown 模板路径",
    )
    parser.add_argument("--summary-output", default=None, help="输出发布判断摘要 Markdown 文件路径")
    parser.add_argument(
        "--summary-template",
        default=str(DEFAULT_SUMMARY_TEMPLATE),
        help="发布判断摘要模板路径",
    )
    parser.add_argument(
        "--summary-only",
        action="store_true",
        help="仅生成摘要，不生成全量报告",
    )
    return parser.parse_args()


def load_workbook(input_path: Path):
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    try:
        return openpyxl.load_workbook(input_path, data_only=True)
    except Exception as exc:  # pragma: no cover - 依赖 openpyxl 具体异常类型
        raise RuntimeError(f"无法读取工作簿: {input_path}") from exc


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_metric_key(value: object) -> str:
    return normalize_text(value).upper()


def detect_overview_sheet(workbook) -> str | None:
    for sheet in workbook.worksheets:
        if "测试报告" in sheet.title:
            return sheet.title
    return None


def count_status_words(sheet) -> dict[str, int]:
    counts = {"pass": 0, "fail": 0, "notest": 0, "block": 0}

    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            normalized = normalize_metric_key(cell)
            status_key = STATUS_VALUE_MAP.get(normalized)
            if status_key:
                counts[status_key] += 1

    return counts


def extract_key_value_metrics(sheet) -> dict[str, int | None]:
    metrics = {value: None for value in METRIC_KEY_MAP.values()}

    for row in sheet.iter_rows(values_only=True):
        for index in range(len(row) - 1):
            key = normalize_metric_key(row[index])
            metric_name = METRIC_KEY_MAP.get(key)
            metric_value = row[index + 1]
            if metric_name is None or not isinstance(metric_value, (int, float)):
                continue
            metrics[metric_name] = int(metric_value)

    return metrics


def fallback_metrics(workbook) -> dict[str, int | None]:
    merged = {
        "total_cases": 0,
        "pass": 0,
        "fail": 0,
        "notest": 0,
        "block": 0,
        "di": None,
        "serious_issues": None,
    }

    for sheet in workbook.worksheets:
        counts = count_status_words(sheet)
        merged["pass"] += counts["pass"]
        merged["fail"] += counts["fail"]
        merged["notest"] += counts["notest"]
        merged["block"] += counts["block"]

    merged["total_cases"] = (
        merged["pass"] + merged["fail"] + merged["notest"] + merged["block"]
    )
    return merged


def has_complete_overview_metrics(metrics: dict[str, int | None]) -> bool:
    required = ("total_cases", "pass", "fail", "notest", "block")
    return all(metrics.get(key) is not None for key in required)


def infer_sheet_role(sheet_name: str, overview_sheet: str | None) -> str:
    if overview_sheet and sheet_name == overview_sheet:
        return "总表"
    if "专项" in sheet_name:
        return "专项页"
    if "测试" in sheet_name:
        return "功能主表"
    if "WIFI" in sheet_name.upper() or "弱网" in sheet_name:
        return "弱网页"
    if "升级" in sheet_name:
        return "升级页"
    return "分析页"


def summarize_workbook(workbook, overview_sheet: str | None) -> dict:
    sheets: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        sheets.append(
            {
                "name": sheet.title,
                "role": infer_sheet_role(sheet.title, overview_sheet),
                "rows": sheet.max_row,
                "cols": sheet.max_column,
            }
        )
    return {"sheet_count": len(sheets), "sheets": sheets}


def summarize_sheet_findings(sheet) -> str:
    counts = count_status_words(sheet)
    risk_count = counts["fail"] + counts["notest"] + counts["block"]
    if risk_count == 0:
        return "未识别到高风险状态。"
    return (
        f"发现 {risk_count} 条高风险状态，"
        f"其中 FAIL={counts['fail']}、NT={counts['notest']}、Block={counts['block']}。"
    )


def build_risk_score(counts: dict[str, int]) -> int:
    return counts["fail"] * 20 + counts["block"] * 20 + counts["notest"]


def has_meaningful_risk(counts: dict[str, int]) -> bool:
    return counts["fail"] > 0 or counts["block"] > 0 or counts["notest"] > 0


def collect_module_findings(workbook, overview_sheet: str | None) -> list[dict[str, str]]:
    findings: list[tuple[int, dict[str, str]]] = []

    for sheet in workbook.worksheets:
        if overview_sheet and sheet.title == overview_sheet:
            continue

        counts = count_status_words(sheet)
        if not has_meaningful_risk(counts):
            continue

        findings.append(
            (
                build_risk_score(counts),
                {
                    "sheet": sheet.title,
                    "summary": summarize_sheet_findings(sheet),
                },
            )
        )

    findings.sort(key=lambda item: item[0], reverse=True)
    return [item[1] for item in findings[:5]]


def extract_risk_rows(sheet) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    for row in sheet.iter_rows(values_only=True):
        status = None
        text_candidates: list[str] = []

        for cell in row:
            value = normalize_text(cell)
            if not value:
                continue

            status_key = STATUS_VALUE_MAP.get(normalize_metric_key(value))
            if status_key:
                status = status_key
                continue

            text_candidates.append(value)

        if status not in {"fail", "notest", "block"} or not text_candidates:
            continue

        issue_text = max(text_candidates, key=len)
        rows.append(
            {
                "sheet": sheet.title,
                "status": status,
                "text": issue_text,
            }
        )

    return rows


def classify_risk_text(text: str) -> str | None:
    normalized = normalize_text(text)
    if not normalized:
        return None

    for label in RISK_PRIORITY:
        for keyword in RISK_KEYWORDS[label]:
            if keyword in normalized:
                return label
    return None


def build_risk_cluster_details(workbook, overview_sheet: str | None) -> list[dict[str, object]]:
    buckets: dict[str, dict[str, object]] = {}

    for sheet in workbook.worksheets:
        if overview_sheet and sheet.title == overview_sheet:
            continue

        sheet_score = build_risk_score(count_status_words(sheet))
        for item in extract_risk_rows(sheet):
            label = classify_risk_text(item["text"])
            if label is None:
                continue

            bucket = buckets.setdefault(
                label,
                {
                    "label": label,
                    "count": 0,
                    "examples": [],
                    "source_sheets": set(),
                    "_score": 0,
                },
            )
            bucket["count"] += 1
            if item["text"] not in bucket["examples"] and len(bucket["examples"]) < 3:
                bucket["examples"].append(item["text"])
            if item["sheet"] not in bucket["source_sheets"]:
                bucket["_score"] += sheet_score
                bucket["source_sheets"].add(item["sheet"])

    priority_index = {label: index for index, label in enumerate(RISK_PRIORITY)}
    ordered = sorted(
        buckets.values(),
        key=lambda item: (
            -int(item["_score"]),
            -int(item["count"]),
            priority_index.get(str(item["label"]), len(RISK_PRIORITY)),
        ),
    )

    details: list[dict[str, object]] = []
    for item in ordered:
        details.append(
            {
                "label": item["label"],
                "count": item["count"],
                "examples": item["examples"],
                "source_sheets": sorted(item["source_sheets"]),
            }
        )

    return details


def categorize_risk(sheet_name: str) -> str:
    upper_name = sheet_name.upper()
    if "音频" in sheet_name:
        return "音频"
    if "预览" in sheet_name or "直播" in sheet_name:
        return "预览传输"
    if "回放" in sheet_name or "TF" in upper_name or "存储" in sheet_name:
        return "存储回放"
    if "WIFI" in upper_name or "弱网" in sheet_name:
        return "弱网"
    if "升级" in sheet_name:
        return "升级稳定性"
    return sheet_name


def build_risk_clusters(workbook, overview_sheet: str | None) -> list[str]:
    details = build_risk_cluster_details(workbook, overview_sheet)
    if details:
        return [str(item["label"]) for item in details[:5]]

    candidates: list[tuple[int, str]] = []

    for sheet in workbook.worksheets:
        if overview_sheet and sheet.title == overview_sheet:
            continue

        counts = count_status_words(sheet)
        if counts["fail"] == 0 and counts["block"] == 0:
            continue

        label = categorize_risk(sheet.title)
        candidates.append((build_risk_score(counts), label))

    candidates.sort(key=lambda item: item[0], reverse=True)

    clusters: list[str] = []
    for _, label in candidates:
        if label not in clusters:
            clusters.append(label)
        if len(clusters) == 5:
            break

    return clusters


def build_retest_suggestions(risk_clusters: list[str]) -> list[str]:
    mapping = {
        "音频": "复测音频：底噪/回声/啸叫/吞字/连续性",
        "预览传输": "复测预览：弱网/长时预览/帧率与延迟/同步性",
        "预览稳定性": "复测预览：弱网/长时预览/帧率与延迟/同步性",
        "存储回放": "复测存储：TF 卡兼容/卡录首检/回放稳定/文件可用性",
        "弱网": "复测网络：穿墙/丢包/重连/码率自适应",
        "升级稳定性": "复测升级：升级成功率/断电恢复/版本回滚",
    }
    result: list[str] = []
    for label in risk_clusters[:5]:
        result.append(mapping.get(label, f"复测模块：{label}（优先复核 FAIL/Block 用例）"))
    return result


def get_platform_semantic_profile(risk_label: str) -> dict[str, object] | None:
    profile = PLATFORM_RISK_PROFILES.get(risk_label)
    if profile is None:
        return None
    return {**PLATFORM_SEMANTIC_DEFAULTS, **profile}


def build_platform_semantic_mapping(risk_clusters: list[str]) -> list[dict[str, object]]:
    mappings: list[dict[str, object]] = []
    for label in risk_clusters[:5]:
        semantic = get_platform_semantic_profile(label)
        if semantic is None:
            continue
        mappings.append({"risk_label": label, **semantic})
    return mappings


def build_release_judgment(metrics: dict[str, int | None]) -> dict[str, str]:
    threshold = "DI <= 12 且 致命+严重 <= 2"
    di = metrics.get("di")
    serious = metrics.get("serious_issues")

    if di is None or serious is None:
        return {
            "decision": "需人工判断",
            "threshold": threshold,
            "gap": "缺少 DI 或严重问题数，需结合人工复核。",
        }

    if di <= 12 and serious <= 2:
        return {
            "decision": "建议发布",
            "threshold": threshold,
            "gap": "当前指标已满足发布门槛。",
        }

    return {
        "decision": "不建议发布",
        "threshold": threshold,
        "gap": f"DI={di}，严重问题数={serious}，均未满足门槛。",
    }


def extract_basic_info(overview_sheet, input_path: Path) -> dict[str, str]:
    basic_info = {"项目": input_path.stem}
    if overview_sheet is None:
        return basic_info

    mapping = {
        "项目": "项目",
        "设备型号": "设备型号",
        "固件版本": "固件版本",
        "APP版本": "APP版本",
        "APP": "APP版本",
        "测试时间": "测试时间",
        "测试完成时间": "测试时间",
        "测试人员": "测试人员",
    }

    for row in overview_sheet.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        key = normalize_text(row[0])
        if key in mapping and row[1] is not None:
            basic_info[mapping[key]] = normalize_text(row[1])

    return basic_info


def build_final_conclusion(
    judgment: dict[str, str], risk_clusters: list[str], used_fallback: bool
) -> str:
    if judgment["decision"] == "建议发布":
        prefix = "当前版本核心指标满足发布门槛，可进入后续发布确认。"
    elif judgment["decision"] == "不建议发布":
        prefix = "当前版本仍存在明显风险，不建议直接发布。"
    else:
        prefix = "当前版本缺少完整总表指标，需结合人工复核后再做发布判断。"

    if risk_clusters:
        prefix += f" 主要风险集中在：{'、'.join(risk_clusters)}。"
    if used_fallback:
        prefix += " 本次结论包含降级统计结果。"
    return prefix


def extract_report_context(input_path: Path) -> dict:
    workbook = load_workbook(input_path)
    overview_name = detect_overview_sheet(workbook)
    overview_sheet = workbook[overview_name] if overview_name else None

    metrics = extract_key_value_metrics(overview_sheet) if overview_sheet else None
    used_fallback = metrics is None or not has_complete_overview_metrics(metrics)
    overall_metrics = fallback_metrics(workbook) if used_fallback else metrics

    workbook_summary = summarize_workbook(workbook, overview_name)
    risk_cluster_details = build_risk_cluster_details(workbook, overview_name)
    risk_clusters = (
        [str(item["label"]) for item in risk_cluster_details[:5]]
        if risk_cluster_details
        else build_risk_clusters(workbook, overview_name)
    )
    release_judgment = build_release_judgment(overall_metrics)

    return {
        "title": f"{input_path.stem} 测试报告学习摘要",
        "source": str(input_path),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "overview_sheet": overview_name,
        "used_fallback": used_fallback,
        "basic_info": extract_basic_info(overview_sheet, input_path),
        "workbook_summary": workbook_summary,
        "overall_metrics": overall_metrics,
        "release_judgment": release_judgment,
        "module_findings": collect_module_findings(workbook, overview_name),
        "risk_cluster_details": risk_cluster_details,
        "risk_clusters": risk_clusters,
        "platform_semantic_mapping": build_platform_semantic_mapping(risk_clusters),
        "final_conclusion": build_final_conclusion(
            release_judgment, risk_clusters, used_fallback
        ),
    }


def format_basic_info_lines(basic_info: dict[str, object]) -> str:
    return "\n".join(f"- {key}: {value}" for key, value in basic_info.items())


def format_workbook_summary_lines(workbook_summary: dict[str, object]) -> str:
    lines: list[str] = []
    for item in workbook_summary.get("sheets", []):
        name = item.get("name", "未知工作表")
        role = item.get("role", "未分类")
        rows = item.get("rows")
        cols = item.get("cols")
        if rows is not None and cols is not None:
            lines.append(f"- {name}: {role}，{rows} 行，{cols} 列")
        else:
            lines.append(f"- {name}: {role}")
    return "\n".join(lines)


def format_overall_metrics_lines(overall_metrics: dict[str, object]) -> str:
    labels = [
        ("总用例", overall_metrics.get("total_cases")),
        ("Pass", overall_metrics.get("pass")),
        ("Fail", overall_metrics.get("fail")),
        ("NoTest", overall_metrics.get("notest")),
        ("Block", overall_metrics.get("block")),
        ("DI", overall_metrics.get("di")),
        ("严重问题数", overall_metrics.get("serious_issues")),
    ]
    return "\n".join(f"- {label}: {value}" for label, value in labels)


def format_module_findings_lines(module_findings: list[dict[str, str]]) -> str:
    if not module_findings:
        return "- 未提取到重点模块结论"
    return "\n".join(
        f"- {item['sheet']}: {item['summary']}" for item in module_findings
    )


def format_risk_cluster_lines(risk_clusters: list[str]) -> str:
    if not risk_clusters:
        return "- 未识别到明显风险聚类"
    return "\n".join(f"- {item}" for item in risk_clusters)


def format_platform_mapping_lines(platform_mapping: list[dict[str, object]]) -> str:
    if not platform_mapping:
        return "- 未建立平台语义映射"

    lines: list[str] = []
    for item in platform_mapping:
        traits = " / ".join(str(value) for value in item["integration_traits"])
        surfaces = " / ".join(str(value) for value in item["observation_surfaces"])
        lines.append(
            "- "
            f"{item['risk_label']}: "
            f"映射到 `{item['ha_domain']}` 域的 `{item['entity_scope']}` 观察面；"
            f"集成特征={traits}；"
            f"重点字段={surfaces}；"
            f"{item['diagnostic_focus']}"
        )
    return "\n".join(lines)


def format_platform_impact_lines(platform_mapping: list[dict[str, object]]) -> str:
    if not platform_mapping:
        return "- 未识别平台影响面"

    lines: list[str] = []
    for item in platform_mapping[:3]:
        traits = " / ".join(str(value) for value in item["integration_traits"])
        surfaces = " / ".join(str(value) for value in item["observation_surfaces"])
        lines.append(
            "- "
            f"{item['risk_label']}: "
            f"`{item['ha_domain']}`.{item['entity_scope']} "
            f"(特征: {traits}; 观察面: {surfaces})"
        )
    return "\n".join(lines)


def format_core_metrics_lines(overall_metrics: dict[str, object]) -> str:
    return format_overall_metrics_lines(overall_metrics)


def format_top_risks_lines(risk_clusters: list[str]) -> str:
    if not risk_clusters:
        return "- 未识别到明显风险"
    return "\n".join(f"- {item}" for item in risk_clusters[:5])


def format_blockers_lines(context: dict) -> str:
    decision = context["release_judgment"]["decision"]
    if decision == "建议发布":
        return "- 无明显阻塞项"

    lines = [f"- {context['release_judgment']['gap']}"]
    for item in context.get("module_findings", [])[:3]:
        lines.append(f"- {item['sheet']}: {item['summary']}")
    return "\n".join(lines)


def render_release_summary(context: dict, template_path: Path | None = None) -> str:
    resolved_template = template_path or DEFAULT_SUMMARY_TEMPLATE
    if not resolved_template.exists():
        raise FileNotFoundError(f"摘要模板不存在: {resolved_template}")

    template = resolved_template.read_text(encoding="utf-8")
    retest_suggestions = build_retest_suggestions(context.get("risk_clusters", []))
    retest_suggestions_lines = "\n".join(f"- {item}" for item in retest_suggestions)

    return template.format(
        title=context["title"],
        source=context["source"],
        date=context["date"],
        release_decision=context["release_judgment"]["decision"],
        release_threshold=context["release_judgment"]["threshold"],
        release_gap=context["release_judgment"]["gap"],
        core_metrics_lines=format_core_metrics_lines(context["overall_metrics"]),
        top_risks_lines=format_top_risks_lines(context.get("risk_clusters", [])),
        blockers_lines=format_blockers_lines(context),
        platform_impact_lines=format_platform_impact_lines(
            context.get("platform_semantic_mapping", [])
        ),
        retest_suggestions_lines=retest_suggestions_lines,
    )


def render_report(context: dict, template_path: Path | None = None) -> str:
    resolved_template = template_path or DEFAULT_TEMPLATE
    template = resolved_template.read_text(encoding="utf-8")

    fallback_notice = ""
    if context.get("used_fallback"):
        fallback_notice = "> 注：未识别标准总表，以下结论基于全工作簿状态词降级统计。\n"

    return template.format(
        title=context["title"],
        source=context["source"],
        date=context["date"],
        fallback_notice=fallback_notice,
        basic_info_lines=format_basic_info_lines(context["basic_info"]),
        workbook_summary_lines=format_workbook_summary_lines(context["workbook_summary"]),
        overall_metrics_lines=format_overall_metrics_lines(context["overall_metrics"]),
        release_decision=context["release_judgment"]["decision"],
        release_threshold=context["release_judgment"]["threshold"],
        release_gap=context["release_judgment"]["gap"],
        module_findings_lines=format_module_findings_lines(context["module_findings"]),
        risk_clusters_lines=format_risk_cluster_lines(context["risk_clusters"]),
        platform_mapping_lines=format_platform_mapping_lines(
            context.get("platform_semantic_mapping", [])
        ),
        final_conclusion=context["final_conclusion"],
    )


def main() -> int:
    configure_stdio()
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    template_path = Path(args.template)
    summary_output = Path(args.summary_output) if args.summary_output else None
    summary_template = Path(args.summary_template) if args.summary_template else None

    try:
        if args.summary_only and summary_output is None:
            raise ValueError("--summary-only 需要与 --summary-output 一起使用")

        context = extract_report_context(input_path)

        if not args.summary_only:
            markdown = render_report(context, template_path=template_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(markdown, encoding="utf-8")
            print(f"已生成报告: {output_path}")

        if summary_output is not None:
            summary_md = render_release_summary(context, template_path=summary_template)
            summary_output.parent.mkdir(parents=True, exist_ok=True)
            summary_output.write_text(summary_md, encoding="utf-8")
            print(f"已生成摘要: {summary_output}")

        return 0
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
