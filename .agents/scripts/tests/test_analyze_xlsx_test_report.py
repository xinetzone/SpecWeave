"""analyze-xlsx-test-report.py 的失败测试。"""

import importlib.util
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "analyze-xlsx-test-report.py"


def load_report_module():
    """按未来脚本路径动态加载模块，当前阶段应因脚本缺失而失败。"""
    spec = importlib.util.spec_from_file_location("analyze_xlsx_test_report", SCRIPT_PATH)
    if spec is None or spec.loader is None or not SCRIPT_PATH.exists():
        raise FileNotFoundError(f"脚本不存在: {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def build_standard_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试报告"
    ws.append(["项目", "IPC-232"])
    ws.append(["设备型号", "单目1M插值3M232"])
    ws.append(["固件版本", "1.0.0"])
    ws.append(["APP版本", "6.0.0"])
    ws.append(["总用例", 10])
    ws.append(["Pass", 7])
    ws.append(["Fail", 2])
    ws.append(["NoTest", 1])
    ws.append(["Block", 0])
    ws.append(["DI", 14])
    ws.append(["严重问题数", 3])

    audio = wb.create_sheet("音频专项")
    audio.append(["模块", "现象", "状态"])
    audio.append(["音频", "底噪明显", "FAIL"])
    audio.append(["音频", "回声轻微", "NT"])

    preview = wb.create_sheet("预览稳定性")
    preview.append(["模块", "现象", "状态"])
    preview.append(["预览", "长时预览正常", "PASS"])
    preview.append(["预览", "弱网卡顿", "FAIL"])

    wb.save(path)
    return path


def build_realistic_overview_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试报告"
    ws.append(["232单目1M插值3M测试报告"])
    ws.append(["基本信息"])
    ws.append(["设备型号", "232单目"])
    ws.append(["固件版本", "[仅供测试]demo.zip"])
    ws.append(["APP", "小强当家5.0.0.65"])
    ws.append(["测试人员", "李翠"])
    ws.append(["测试完成时间", "2026-04-02"])
    ws.append(["测试结果", "新增用例", "", "所有用例", 489])
    ws.append(["", "Pass", "", "Pass", 380])
    ws.append(["", "Fail", "", "Fail", 26])
    ws.append(["", "NoTest", "", "NoTest", 70])
    ws.append(["", "Block", "", "Block", 13])
    ws.append(["", "新增致命问题", "", "所有致命问题", 0])
    ws.append(["", "新增严重问题", "", "所有严重问题", 11])
    ws.append(["", "最终结果", "", "DI值", 35])

    audio = wb.create_sheet("06音频专项测试")
    audio.append(["模块", "现象", "状态"])
    audio.append(["音频", "底噪明显", "FAIL"])
    audio.append(["音频", "回声轻微", "FAIL"])

    wifi = wb.create_sheet("WiFi穿墙测试")
    wifi.append(["点位", "现象", "状态"])
    wifi.append(["点7", "拉流卡顿", "FAIL"])

    wb.save(path)
    return path


def test_extract_report_context_with_standard_overview(tmp_path) -> None:
    report = load_report_module()
    workbook_path = build_standard_workbook(tmp_path / "sample.xlsx")

    context = report.extract_report_context(workbook_path)

    assert context["overview_sheet"] == "测试报告"
    assert context["overall_metrics"]["total_cases"] == 10
    assert context["overall_metrics"]["pass"] == 7
    assert context["overall_metrics"]["fail"] == 2
    assert context["overall_metrics"]["notest"] == 1
    assert context["overall_metrics"]["block"] == 0
    assert context["overall_metrics"]["di"] == 14
    assert context["overall_metrics"]["serious_issues"] == 3


def test_extract_report_context_with_realistic_overview_layout(tmp_path) -> None:
    report = load_report_module()
    workbook_path = build_realistic_overview_workbook(tmp_path / "realistic.xlsx")

    context = report.extract_report_context(workbook_path)

    assert context["used_fallback"] is False
    assert context["overall_metrics"]["total_cases"] == 489
    assert context["overall_metrics"]["pass"] == 380
    assert context["overall_metrics"]["fail"] == 26
    assert context["overall_metrics"]["notest"] == 70
    assert context["overall_metrics"]["block"] == 13
    assert context["overall_metrics"]["di"] == 35
    assert context["overall_metrics"]["serious_issues"] == 11
    assert context["release_judgment"]["decision"] == "不建议发布"
    assert "risk_cluster_details" in context
    assert context["risk_cluster_details"]
    assert context["risk_clusters"][0] in {"弱网", "音频", "预览稳定性"}
    first_detail = context["risk_cluster_details"][0]
    assert set(first_detail.keys()) == {"label", "count", "examples", "source_sheets"}


def test_extract_report_context_prefers_text_risk_clustering(tmp_path) -> None:
    report = load_report_module()
    workbook_path = tmp_path / "text-priority.xlsx"

    wb = openpyxl.Workbook()
    overview = wb.active
    overview.title = "测试报告"
    overview.append(["总用例", 4])
    overview.append(["Pass", 0])
    overview.append(["Fail", 4])
    overview.append(["NoTest", 0])
    overview.append(["Block", 0])

    generic = wb.create_sheet("01功能测试")
    generic.append(["模块", "现象", "状态"])
    generic.append(["模块A", "异常重启后黑屏", "FAIL"])
    generic.append(["模块B", "弱网环境下重连失败", "FAIL"])
    generic.append(["模块C", "底噪明显", "FAIL"])
    generic.append(["模块D", "TF卡录像文件损坏", "FAIL"])
    wb.save(workbook_path)

    context = report.extract_report_context(workbook_path)

    assert context["risk_clusters"][:4] == ["重启恢复", "弱网", "存储回放", "音频"]
    assert context["risk_cluster_details"][0]["label"] == "重启恢复"
    assert "platform_semantic_mapping" in context
    mapping_labels = [item["risk_label"] for item in context["platform_semantic_mapping"]]
    assert "弱网" in mapping_labels
    weak_network_mapping = next(
        item for item in context["platform_semantic_mapping"] if item["risk_label"] == "弱网"
    )
    assert weak_network_mapping["ha_domain"] == "tuya"
    assert "cloud_push" in weak_network_mapping["integration_traits"]
    assert "status" in weak_network_mapping["observation_surfaces"]


def test_extract_report_context_falls_back_to_status_scan(tmp_path) -> None:
    report = load_report_module()
    workbook_path = tmp_path / "fallback.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "功能测试"
    ws.append(["模块", "状态"])
    ws.append(["预览", "PASS"])
    ws.append(["音频", "FAIL"])
    ws.append(["存储", "NT"])
    wb.save(workbook_path)

    context = report.extract_report_context(workbook_path)

    assert context["overview_sheet"] is None
    assert context["used_fallback"] is True
    assert context["overall_metrics"]["pass"] == 1
    assert context["overall_metrics"]["fail"] == 1
    assert context["overall_metrics"]["notest"] == 1


def test_render_report_contains_required_sections() -> None:
    report = load_report_module()
    context = {
        "title": "示例报告",
        "source": "sample.xlsx",
        "date": "2026-07-01",
        "basic_info": {"项目": "IPC-232"},
        "workbook_summary": {"sheet_count": 2, "sheets": [{"name": "测试报告", "role": "总表"}]},
        "overall_metrics": {
            "total_cases": 10,
            "pass": 7,
            "fail": 2,
            "notest": 1,
            "block": 0,
            "di": 14,
            "serious_issues": 3,
        },
        "release_judgment": {
            "decision": "不建议发布",
            "threshold": "DI <= 12 且 致命+严重 <= 2",
            "gap": "DI 和严重问题数均超标",
        },
        "module_findings": [{"sheet": "音频专项", "summary": "音频风险集中在底噪与回声"}],
        "risk_clusters": ["音频", "预览传输"],
        "platform_semantic_mapping": [
            {
                "risk_label": "音频",
                "ha_domain": "tuya",
                "entity_scope": "camera",
                "integration_traits": ["hub", "cloud_push"],
                "observation_surfaces": ["status", "function", "status_range"],
                "diagnostic_focus": "结合实体状态与诊断字段复核音视频能力相关风险。",
            }
        ],
        "final_conclusion": "当前版本不满足发布门槛。",
        "used_fallback": False,
    }

    markdown = report.render_report(context)

    assert "## 结论摘要" in markdown
    assert "## 基本信息" in markdown
    assert "## 工作簿结构" in markdown
    assert "## 总体结果" in markdown
    assert "## 风险聚类" in markdown
    assert "## 平台语义映射" in markdown
    assert "## 最终判断" in markdown


def test_cli_writes_markdown_output(tmp_path) -> None:
    workbook_path = build_standard_workbook(tmp_path / "cli.xlsx")
    output_path = tmp_path / "report.md"

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT_PATH),
            "--input",
            str(workbook_path),
            "--output",
            str(output_path),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )

    assert result.returncode == 0
    assert output_path.exists()
    content = output_path.read_text(encoding="utf-8")
    assert "示例报告" not in content
    assert "## 最终判断" in content


def test_extract_report_context_raises_for_missing_file(tmp_path) -> None:
    report = load_report_module()

    with pytest.raises(FileNotFoundError):
        report.extract_report_context(tmp_path / "missing.xlsx")


def test_cli_returns_nonzero_for_invalid_workbook(tmp_path) -> None:
    workbook_path = tmp_path / "broken.xlsx"
    workbook_path.write_text("not an xlsx", encoding="utf-8")
    output_path = tmp_path / "broken.md"

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT_PATH),
            "--input",
            str(workbook_path),
            "--output",
            str(output_path),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )

    assert result.returncode == 1
    assert "无法读取工作簿" in result.stderr
    assert not output_path.exists()


def test_render_release_summary_function_is_exposed() -> None:
    report = load_report_module()
    assert hasattr(report, "render_release_summary"), "功能缺失: render_release_summary 未实现"


def test_render_release_summary_contains_platform_impact() -> None:
    report = load_report_module()
    context = {
        "title": "发布摘要示例",
        "source": "sample.xlsx",
        "date": "2026-07-01",
        "overall_metrics": {
            "total_cases": 10,
            "pass": 7,
            "fail": 2,
            "notest": 1,
            "block": 0,
            "di": 14,
            "serious_issues": 3,
        },
        "release_judgment": {
            "decision": "不建议发布",
            "threshold": "DI <= 12 且 致命+严重 <= 2",
            "gap": "DI 和严重问题数均超标",
        },
        "risk_clusters": ["弱网", "音频"],
        "platform_semantic_mapping": [
            {
                "risk_label": "弱网",
                "ha_domain": "tuya",
                "entity_scope": "camera",
                "integration_traits": ["hub", "cloud_push"],
                "observation_surfaces": ["status", "function", "diagnostics"],
                "diagnostic_focus": "结合实体状态刷新延迟与诊断字段复核弱网风险。",
            },
            {
                "risk_label": "音频",
                "ha_domain": "tuya",
                "entity_scope": "camera",
                "integration_traits": ["hub", "cloud_push"],
                "observation_surfaces": ["status", "function", "status_range"],
                "diagnostic_focus": "复核音视频能力字段是否存在异常状态。",
            },
        ],
    }

    markdown = report.render_release_summary(context)

    assert "## 平台影响面" in markdown
    assert "弱网" in markdown
    assert "cloud_push" in markdown
    assert "status / function / diagnostics" in markdown


def test_get_platform_semantic_profile_merges_defaults() -> None:
    report = load_report_module()

    profile = report.get_platform_semantic_profile("弱网")

    assert profile is not None
    assert profile["ha_domain"] == "tuya"
    assert profile["entity_scope"] == "camera"
    assert profile["integration_traits"] == ("hub", "cloud_push")
    assert "diagnostics" in profile["observation_surfaces"]


def test_render_release_summary_with_realistic_overview_layout(tmp_path) -> None:
    report = load_report_module()
    workbook_path = build_realistic_overview_workbook(tmp_path / "realistic-summary.xlsx")

    context = report.extract_report_context(workbook_path)
    markdown = report.render_release_summary(context)

    assert context["release_judgment"]["decision"] == "不建议发布"
    assert context["release_judgment"]["gap"] == "DI=35，严重问题数=11，均未满足门槛。"
    assert context["risk_clusters"][:2] == ["音频", "预览稳定性"]
    assert "## 平台影响面" in markdown
    assert "音频: `tuya`.camera" in markdown
    assert "预览稳定性: `tuya`.camera" in markdown
    assert "status / function / status_range" in markdown
    assert "DI=35，严重问题数=11，均未满足门槛。" in markdown
    assert "06音频专项测试: 发现 2 条高风险状态，其中 FAIL=2、NT=0、Block=0。" in markdown
    assert "复测音频：底噪/回声/啸叫/吞字/连续性" in markdown
    assert "复测预览：弱网/长时预览/帧率与延迟/同步性" in markdown


def test_cli_supports_summary_output_and_summary_only_flags(tmp_path) -> None:
    workbook_path = build_standard_workbook(tmp_path / "cli-summary.xlsx")
    output_path = tmp_path / "full.md"
    summary_output = tmp_path / "summary.md"

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT_PATH),
            "--input",
            str(workbook_path),
            "--output",
            str(output_path),
            "--summary-output",
            str(summary_output),
            "--summary-only",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )

    assert (
        result.returncode == 0
    ), f"参数未实现: CLI 未支持 --summary-output/--summary-only (stderr={result.stderr.strip()})"
    assert summary_output.exists()
    assert not output_path.exists()


def test_classify_risk_text_by_keywords() -> None:
    report = load_report_module()

    assert report.classify_risk_text("底噪明显") == "音频"
    assert report.classify_risk_text("拉流卡顿，画面延迟") == "预览稳定性"
    assert report.classify_risk_text("异常重启后黑屏") == "重启恢复"
    assert report.classify_risk_text("TF卡录像文件损坏") == "存储回放"
    assert report.classify_risk_text("弱网环境下重连失败") == "弱网"
    assert report.classify_risk_text("升级后版本回退") == "升级稳定性"
    assert report.classify_risk_text("普通功能验证通过") is None


def test_build_risk_cluster_details_aggregates_examples_and_sheets(tmp_path) -> None:
    report = load_report_module()
    workbook_path = tmp_path / "semantic.xlsx"

    wb = openpyxl.Workbook()
    overview = wb.active
    overview.title = "测试报告"
    overview.append(["总用例", 6])
    overview.append(["Pass", 1])
    overview.append(["Fail", 4])
    overview.append(["NoTest", 1])
    overview.append(["Block", 0])

    audio = wb.create_sheet("06音频专项测试")
    audio.append(["模块", "现象", "状态"])
    audio.append(["音频", "底噪明显", "FAIL"])
    audio.append(["音频", "回声轻微", "NT"])

    wifi = wb.create_sheet("WiFi穿墙测试")
    wifi.append(["点位", "现象", "状态"])
    wifi.append(["点7", "弱网环境下重连失败", "FAIL"])
    wifi.append(["点8", "拉流卡顿", "FAIL"])

    wb.save(workbook_path)

    workbook = report.load_workbook(workbook_path)
    details = report.build_risk_cluster_details(workbook, "测试报告")

    assert details[0]["label"] == "弱网"
    assert details[0]["count"] == 1
    assert "WiFi穿墙测试" in details[0]["source_sheets"]
    assert "弱网环境下重连失败" in details[0]["examples"]

    labels = [item["label"] for item in details]
    assert "音频" in labels
    assert "预览稳定性" in labels
